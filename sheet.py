#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

class Sheet():
	def __init__(self, _file):
		self.instance = load_workbook(_file)
		self.columns = dict()
		
	def load_sheet(self, sheet):												#carrega a sheet   H
		self.sheet = self.instance.get_sheet_by_name(sheet)	

	def load_columns(self, header=1):											#carrega as colunas da planilha   H
		for col_index in range(self.sheet.max_column):
			self.columns[self.sheet.cell(row=header,  column=col_index + 1).value] = col_index			
					
	def search(self, value, column):											#procura o Valor SH
		for row_index in range(self.sheet.max_row):
			if self.sheet.cell(row=row_index+1 , column=column+1).value == value:
				return self.sheet.cell(row=row_index+1 , column=column+1)
		return None

	def add_element(self, columns, value):
		maximo = self.sheet.max_row
		j = 0
		for i in columns:
			self.sheet.cell(row=maximo+1, column=i+1).value = value[j]
			j += 1

	def save(self, name):
		self.instance.save(name)		

