#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from sheet_to_sheet import *
from sheet import *

def main():
	inventarioRjSP = Sheet('./files/Controle de servidores VM_NEW Versao atualizada.xlsx')
	myProducts = Sheet('./files/dell.xlsx')
	corresp = Sheet_to_sheet()

	#carregando sheets
	inventarioRjSP.load_sheet('inventario rio e sp')
	myProducts.load_sheet('MyProductList (002)')

	#carregando cabeçalhos
	inventarioRjSP.load_columns(4)
	myProducts.load_columns()

	#iterando sobre dell e checando
	for row_index in range(myProducts.sheet.max_row):
		if inventarioRjSP.search(myProducts.sheet.cell(row=row_index+1, column=myProducts.columns[corresp.inv_to_dell['SERIAL']]+1).value, inventarioRjSP.columns['SERIAL']) == None:

			#instanciando valores
			serial_value = myProducts.sheet.cell(row=row_index+1, column=myProducts.columns[corresp.inv_to_dell['SERIAL']]+1).value
			modelo_value = myProducts.sheet.cell(row=row_index+1, column=myProducts.columns[corresp.inv_to_dell['MODELO']]+1).value
			hostname_value = myProducts.sheet.cell(row=row_index+1, column=myProducts.columns[corresp.inv_to_dell['HOSTNAME']]+1).value
			start_value = myProducts.sheet.cell(row=row_index+1, column=myProducts.columns[corresp.inv_to_dell['Start Date']]+1).value
			end_value = myProducts.sheet.cell(row=row_index+1, column=myProducts.columns[corresp.inv_to_dell['End Date']]+1).value

			#adicionando serial
			inventarioRjSP.add_element((inventarioRjSP.columns['SERIAL'], inventarioRjSP.columns['MODELO'], inventarioRjSP.columns['HOSTNAME'], inventarioRjSP.columns['FABRICANTE'], inventarioRjSP.columns['Start Date'], inventarioRjSP.columns['End Date']), (serial_value, modelo_value, hostname_value, 'DELL', start_value, end_value))


	#salvando arquivo		
	inventarioRjSP.save('./files/teste_main.xlsx')


if __name__ == '__main__':
	main()