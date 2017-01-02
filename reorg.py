from openpyxl import load_workbook
from sheet import *

def cadastro_valores(_object, start_row, column):
	elements = []
	for i in range(start_row, _object.sheet.max_row):
		if _object.sheet.cell(row=i, column=column).value not in  

def main():
	#intanciando planilha
	inventarioRjSP = Sheet('./files/Controle de servidores VM_NEW Versao atualizada.xlsx')

	#carregando sheet
	inventarioRjSP.load_sheet('inventario rio e sp')

	#carregando cabeçalhos
	inventarioRjSP.load_columns()





