from openpyxl import load_workbook
import glob

'''
class Elemento():
	def __init__(self):
		_colunms = {categoria: , responsavel: , serial: , fabricante: , modelo: , localizacao: , rack: ,
		 patrimonio: , hostname: , em_uso: , said: , contrato: , start_date: , end_date: , legado: }
'''

_file = './files/Controle de servidores VM_NEW Versao atualizada.xlsx'

wb = load_workbook(_file)
sheet = wb.get_sheet_by_name('inventario rio e sp')

for row_index in range(sheet.get_highest_row()):
	print sheet.cell(row=row_index + 1, column=3).value

 